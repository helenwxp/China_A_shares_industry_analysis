# %%
from bs4 import BeautifulSoup
import requests as reqs
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re
from tqdm import tqdm

# %% visit website: zhengce.chinabaogao.com
header = {
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) \
        AppleWebKit/537.36 (KHTML, like Gecko) \
        Chrome/92.0.4515.131 Safari/537.36'}

url = 'http://zhengce.chinabaogao.com/'
resp = reqs.get(url,headers=header)
resp.encoding = 'gb2312'
soup = BeautifulSoup(resp.text,'html.parser')
print(soup.prettify())

# %% retrieve initial links of 28 industries
div = soup.find_all(class_='category clearfix')[0]
inds_link = pd.DataFrame(index=np.arange(len(div.select('ul li h2 a'))),\
    columns=['industry','policylink'])

for i, inds in enumerate(div.select('ul li h2 a')):
    inds_link.loc[i,'industry'] = inds.string
    inds_link.loc[i,'policylink'] = inds['href']

inds_link.set_index('industry', inplace=True)

# %% collect all the policy links of 28 policies
wb = Workbook()
for ind in tqdm(inds_link.index):
    ws = wb.create_sheet(ind) # create a worksheet of certain industry
    cols = ['policy','date','link'] # define columns
    ws.append(cols)
    
    # get the initial url of certain industry
    indurl_init = inds_link.loc[ind, 'policylink']
    r = reqs.get(indurl_init, headers=header)
    r.encoding='gb2312'
    soup = BeautifulSoup(r.text, 'html.parser')

    # get all the url of certain industry
    try:
        endpage = soup.find(class_='endpage')['href']
        n_pages = int(re.findall('\d+',endpage)[0])

        indurl_all = [indurl_init+'list_{}.html'.format(i) for i in range(1,n_pages+1)]
    except TypeError:
        indurl_all = [indurl_init]

    for indurl in indurl_all:
        r = reqs.get(indurl, headers=header)
        r.encoding='gb2312'

        soup = BeautifulSoup(r.text, 'html.parser')

        ul = soup.find_all(class_='pagelist')[0]
        for li in ul.select('li'):
            policy_list = []
            try:
                if li.select('h3 a')[0].string == None:
                    policy_list.append(li.select('h3 a')[0]['title']) # 'policy'
                else:
                    policy_list.append(li.select('h3 a')[0].string) # 'policy'
                policy_list.append(li.select('span')[0].string) # 'date'
                policy_list.append(li.select('h3 a')[0]['href']) # 'link'
                ws.append(policy_list)
            except IndexError:
                continue


wb.save('industry_policies.xlsx')
wb.close()

# %% filter policies since 2019 & scrape detailed contents
inds = inds_link.index
wb = load_workbook('industry_policies.xlsx')

for ind in tqdm(inds):
    # policies since 2019
    policies = pd.read_excel('industry_policies.xlsx', sheet_name=ind)
    policies['date'] = pd.to_datetime(policies['date'])
    policies.set_index('date',inplace=True)
    policies = policies[policies.index>='2019/01/01']

    ws = wb[ind]
    ws['D1'] = 'contents'
    for i, link in enumerate(policies.link):
        r = reqs.get(link, headers=header)
        r.encoding = 'gb2312'

        soup = BeautifulSoup(r.text, 'html.parser')
        if soup.find(class_= 'cbg-a-d-content') == None:
            continue
        else:
            texts = soup.find(class_= 'cbg-a-d-content').get_text()
        texts = ''.join(texts.split())

        ws['D{}'.format(i+2)] = texts.strip('中国报告网提示：')

wb.save('industry_policies_with_texts.xlsx')
wb.close()

